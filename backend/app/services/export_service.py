"""
Taurus Vision — Export Service

CSV va Excel formatida professional ma'lumot eksporti.

METODLAR:
    export_animals_csv()    — Jonivorlar CSV (filtrlangan)
    export_animals_excel()  — Jonivorlar Excel (professional, ko'p varaqli)  ← B7 yangi
    export_detections_csv() — Deteksiyalar CSV (sana oralig'i)
    export_weights_excel()  — Og'irlik o'lchovlari Excel (har jonivor uchun varaq)
    export_all_data_excel() — To'liq arxiv Excel (4 varaqli)

TUZATILGAN XATOLAR (oldingi versiyada):
    ✗ Detection.detected_at     → ✓ Detection.timestamp
    ✗ detection.confidence_score → ✓ detection.confidence
    ✗ detection.bbox_x/y/...    → ✓ detection.bbox (JSON dict: x, y, w, h)
    (WeightMeasurement.confidence_score — to'g'ri, o'zgartirilmadi)
"""

from __future__ import annotations

import io
from datetime import date, datetime, timedelta, timezone
from typing import Any, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from sqlalchemy import and_, func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.logging_config import get_logger
from app.models.animal import Animal, AnimalStatus
from app.models.detection import Detection
from app.models.weight_measurement import WeightMeasurement

logger = get_logger(__name__)


# =============================================================================
# OPENPYXL STYLE HELPERS
# =============================================================================

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold: bool = False, size: int = 11, color: str = "000000", italic: bool = False) -> Font:
    return Font(bold=bold, size=size, color=color, italic=italic, name="Calibri")

def _border_thin() -> Border:
    thin = Side(style="thin", color="D1D5DB")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _align(horizontal: str = "left", wrap: bool = False) -> Alignment:
    return Alignment(horizontal=horizontal, vertical="center", wrap_text=wrap)

# Status → hex colour
_STATUS_HEX: dict[str, tuple[str, str]] = {
    "active":      ("D1FAE5", "065F46"),  # emerald bg, dark text
    "sick":        ("FEE2E2", "991B1B"),  # red
    "quarantine":  ("FEF3C7", "92400E"),  # amber
    "sold":        ("DBEAFE", "1E40AF"),  # blue
    "deceased":    ("F3F4F6", "374151"),  # gray
    "transferred": ("EDE9FE", "5B21B6"),  # violet
}

_SPECIES_LABEL = {
    "cattle": "Qoramol",
    "sheep":  "Qo'y",
    "goat":   "Echki",
    "horse":  "Ot",
    "other":  "Boshqa",
}

_GENDER_LABEL = {
    "male":    "Erkak",
    "female":  "Urg'ochi",
    "unknown": "Noma'lum",
}

_STATUS_LABEL = {
    "active":      "Faol",
    "sick":        "Kasal",
    "quarantine":  "Karantin",
    "sold":        "Sotilgan",
    "deceased":    "Vafot etgan",
    "transferred": "Ko'chirilgan",
}


def _set_col_widths(ws: Any, widths: list[int]) -> None:
    """Ustun kengliklarini belgilash."""
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_header_row(ws: Any, headers: list[str], row: int = 1) -> None:
    """Professional sarlavha qatori yozadi."""
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=title)
        cell.font      = _font(bold=True, size=11, color="FFFFFF")
        cell.fill      = _fill("1E3A5F")           # qoʻngʻir ko'k
        cell.alignment = _align("center")
        cell.border    = _border_thin()


def _write_data_cell(
    ws: Any,
    row: int,
    col: int,
    value: Any,
    even_row: bool = False,
    align: str = "left",
    bold: bool = False,
    color: str | None = None,
    bg: str | None = None,
    number_format: str | None = None,
) -> None:
    """Bitta ma'lumot katakchasi yozadi."""
    cell = ws.cell(row=row, column=col, value=value)
    base_bg = "F9FAFB" if even_row else "FFFFFF"
    cell.fill      = _fill(bg or base_bg)
    cell.font      = _font(bold=bold, color=color or "111827")
    cell.alignment = _align(align)
    cell.border    = _border_thin()
    if number_format:
        cell.number_format = number_format


# =============================================================================
# EXPORT SERVICE
# =============================================================================

class ExportService:
    """
    Professional ma'lumot eksport servisi.

    Barcha metodlar async, to'liq type-annotated.
    Xato bo'lsa — log yozadi va qayta raise qiladi.
    """

    # =========================================================================
    # ANIMALS — CSV
    # =========================================================================

    async def export_animals_csv(
        self,
        db: AsyncSession,
        filters: Optional[dict[str, Any]] = None,
    ) -> bytes:
        """
        Jonivorlarni CSV formatida eksport qiladi.

        Args:
            db:      DB sessiyasi
            filters: Ixtiyoriy filtrlar: status, species, gender, tag_id

        Returns:
            UTF-8 kodlangan CSV baytlari

        CSV ustunlari:
            id, tag_id, species, gender, status, breed,
            acquisition_date, total_detections, last_detected_at, notes
        """
        logger.info(f"[export] Animals CSV boshlanmoqda: filters={filters}")

        query = select(Animal)
        if filters:
            conds = []
            if filters.get("status"):
                conds.append(Animal.status == AnimalStatus(filters["status"]))
            if filters.get("species"):
                conds.append(Animal.species == filters["species"])
            if filters.get("gender"):
                conds.append(Animal.gender == filters["gender"])
            if filters.get("tag_id"):
                conds.append(Animal.tag_id.ilike(f"%{filters['tag_id']}%"))
            if conds:
                query = query.where(and_(*conds))

        result  = await db.execute(query.order_by(Animal.tag_id))
        animals = result.scalars().all()

        rows = [
            {
                "id":               a.id,
                "tag_id":           a.tag_id,
                "species":          a.species.value if hasattr(a.species, "value") else str(a.species),
                "gender":           a.gender.value  if hasattr(a.gender,  "value") else str(a.gender),
                "status":           a.status.value  if hasattr(a.status,  "value") else str(a.status),
                "breed":            a.breed or "",
                "acquisition_date": a.acquisition_date.date().isoformat() if a.acquisition_date else "",
                "total_detections": a.total_detections,
                "last_detected_at": a.last_detected_at.isoformat() if a.last_detected_at else "",
                "notes":            a.notes or "",
            }
            for a in animals
        ]

        buf = io.BytesIO()
        pd.DataFrame(rows).to_csv(buf, index=False, encoding="utf-8")
        raw = buf.getvalue()
        buf.close()

        logger.info(f"[export] Animals CSV tayyor: {len(animals)} ta, {len(raw)} bayt")
        return raw

    # =========================================================================
    # ANIMALS — EXCEL (B7 — yangi feature)
    # =========================================================================

    async def export_animals_excel(
        self,
        db: AsyncSession,
        filters: Optional[dict[str, Any]] = None,
    ) -> bytes:
        """
        Jonivorlarni professional Excel formatida eksport qiladi.

        EXCEL TARKIBI:
            Varaq 1 — "Ro'yxat"      : Barcha jonivorlar; holat ranglari, muzlatilgan sarlavha
            Varaq 2 — "Statistika"   : Tur va holat bo'yicha taqsimot; umumiy ko'rsatkichlar

        PROFESSIONAL FORMATLASH:
            - Qoʻyuq ko'k sarlavha satr
            - Holat katakchalari holat rangi bilan (faol=yashil, kasal=qizil …)
            - Jadval kengligi avtomatik moslashtirilgan
            - Muzlatilgan sarlavha (freeze_panes)
            - Sahifa osti va tepasi (print_title)
            - Alternativ qator ranglari
            - Raqam formatlash

        Args:
            db:      DB sessiyasi
            filters: Ixtiyoriy filtrlar: status, species, gender, tag_id

        Returns:
            .xlsx baytlari (openpyxl, engine-siz pandas yo'q)
        """
        logger.info(f"[export] Animals Excel boshlanmoqda: filters={filters}")

        # ------------------------------------------------------------------
        # 1. Ma'lumotlarni yuklash
        # ------------------------------------------------------------------
        query = select(Animal)
        if filters:
            conds = []
            if filters.get("status"):
                conds.append(Animal.status == AnimalStatus(filters["status"]))
            if filters.get("species"):
                conds.append(Animal.species == filters["species"])
            if filters.get("gender"):
                conds.append(Animal.gender == filters["gender"])
            if filters.get("tag_id"):
                conds.append(Animal.tag_id.ilike(f"%{filters['tag_id']}%"))
            if conds:
                query = query.where(and_(*conds))

        result  = await db.execute(query.order_by(Animal.tag_id))
        animals = result.scalars().all()

        # ------------------------------------------------------------------
        # 2. Workbook yaratish
        # ------------------------------------------------------------------
        wb = Workbook()
        wb.remove(wb.active)   # default blank sheet

        now_str = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

        # ==================================================================
        # VARAQ 1 — Ro'yxat
        # ==================================================================
        ws1 = wb.create_sheet("Ro'yxat")

        # Sarlavha bloki
        ws1.merge_cells("A1:J1")
        title_cell = ws1["A1"]
        title_cell.value     = "TAURUS VISION — JONIVORLAR RO'YXATI"
        title_cell.font      = _font(bold=True, size=14, color="FFFFFF")
        title_cell.fill      = _fill("0F172A")
        title_cell.alignment = _align("center")
        ws1.row_dimensions[1].height = 32

        ws1.merge_cells("A2:J2")
        meta_cell = ws1["A2"]
        meta_cell.value     = f"Eksport: {now_str}   |   Jami: {len(animals)} ta jonivor"
        meta_cell.font      = _font(italic=True, size=10, color="6B7280")
        meta_cell.fill      = _fill("F8FAFC")
        meta_cell.alignment = _align("center")
        ws1.row_dimensions[2].height = 20

        # Ustun sarlavhalari (3-qator)
        headers1 = [
            "#", "Tag ID", "Tur", "Jins", "Zot",
            "Holat", "Sotib olingan", "Jami aniqlash",
            "Oxirgi ko'rinish", "Izoh",
        ]
        _write_header_row(ws1, headers1, row=3)
        ws1.row_dimensions[3].height = 22
        ws1.freeze_panes = "A4"     # sarlavhani muzlatish

        # Ma'lumot qatorlari
        for i, a in enumerate(animals, start=1):
            row_num  = i + 3
            even_row = (i % 2 == 0)

            status_val = a.status.value if hasattr(a.status, "value") else str(a.status)
            species_val = a.species.value if hasattr(a.species, "value") else str(a.species)
            gender_val  = a.gender.value  if hasattr(a.gender,  "value") else str(a.gender)

            status_bg, status_fg = _STATUS_HEX.get(status_val, ("F3F4F6", "374151"))

            base_cells = [
                (i,                                  "center", False, None,      None),
                (a.tag_id,                           "left",   True,  "374151",  None),
                (_SPECIES_LABEL.get(species_val, species_val), "left", False, None, None),
                (_GENDER_LABEL.get(gender_val, gender_val),   "left", False, None, None),
                (a.breed or "—",                     "left",   False, "6B7280",  None),
            ]
            for col, (val, align, bold, clr, bg) in enumerate(base_cells, start=1):
                _write_data_cell(ws1, row_num, col, val, even_row, align, bold, clr, bg)

            # Holat katakchasi — rang bilan
            _write_data_cell(
                ws1, row_num, 6,
                _STATUS_LABEL.get(status_val, status_val),
                even_row=False, align="center", bold=True,
                color=status_fg, bg=status_bg,
            )

            acq = a.acquisition_date.date().isoformat() if a.acquisition_date else "—"
            _write_data_cell(ws1, row_num, 7, acq, even_row, "center")

            _write_data_cell(
                ws1, row_num, 8,
                a.total_detections,
                even_row, "center", bold=(a.total_detections > 0),
                color="1E40AF" if a.total_detections > 0 else "9CA3AF",
            )

            last = a.last_detected_at.strftime("%Y-%m-%d %H:%M") if a.last_detected_at else "—"
            _write_data_cell(ws1, row_num, 9, last, even_row, "center", color="6B7280")

            _write_data_cell(ws1, row_num, 10, a.notes or "", even_row, "left", color="6B7280")

            ws1.row_dimensions[row_num].height = 20

        # Ustun kengliklari
        _set_col_widths(ws1, [6, 14, 14, 12, 18, 14, 16, 14, 20, 30])

        # Agar ma'lumot yo'q bo'lsa
        if not animals:
            ws1.merge_cells("A4:J4")
            empty_cell = ws1["A4"]
            empty_cell.value     = "Filtrga mos jonivor topilmadi."
            empty_cell.font      = _font(italic=True, color="9CA3AF")
            empty_cell.alignment = _align("center")

        # ==================================================================
        # VARAQ 2 — Statistika
        # ==================================================================
        ws2 = wb.create_sheet("Statistika")

        ws2.merge_cells("A1:D1")
        ws2["A1"].value     = "STATISTIKA"
        ws2["A1"].font      = _font(bold=True, size=13, color="FFFFFF")
        ws2["A1"].fill      = _fill("0F172A")
        ws2["A1"].alignment = _align("center")
        ws2.row_dimensions[1].height = 28

        # --- Tur bo'yicha ---
        def _stat_header(ws: Any, row: int, text: str) -> None:
            ws.merge_cells(f"A{row}:D{row}")
            c = ws[f"A{row}"]
            c.value     = text
            c.font      = _font(bold=True, size=11, color="FFFFFF")
            c.fill      = _fill("1E3A5F")
            c.alignment = _align("center")
            ws.row_dimensions[row].height = 20

        def _stat_row(ws: Any, row: int, label: str, value: Any, even: bool) -> None:
            for col, (val, align) in enumerate([(label, "left"), (value, "center")], start=1):
                _write_data_cell(ws, row, col, val, even, align)
            # bo'sh D ustunlari
            for col in [3, 4]:
                _write_data_cell(ws, row, col, "", even)

        _stat_header(ws2, 2, "TUR BO'YICHA TAQSIMOT")
        species_counts: dict[str, int] = {}
        for a in animals:
            sv = a.species.value if hasattr(a.species, "value") else str(a.species)
            species_counts[sv] = species_counts.get(sv, 0) + 1

        r = 3
        for i, (sp, cnt) in enumerate(sorted(species_counts.items())):
            _stat_row(ws2, r, _SPECIES_LABEL.get(sp, sp), cnt, i % 2 == 0)
            r += 1

        r += 1
        _stat_header(ws2, r, "HOLAT BO'YICHA TAQSIMOT")
        r += 1
        status_counts: dict[str, int] = {}
        for a in animals:
            sv = a.status.value if hasattr(a.status, "value") else str(a.status)
            status_counts[sv] = status_counts.get(sv, 0) + 1

        for i, (st, cnt) in enumerate(sorted(status_counts.items())):
            bg, fg = _STATUS_HEX.get(st, ("F3F4F6", "374151"))
            cell_l = ws2.cell(row=r, column=1, value=_STATUS_LABEL.get(st, st))
            cell_l.font      = _font(color=fg, bold=True)
            cell_l.fill      = _fill(bg)
            cell_l.alignment = _align()
            cell_l.border    = _border_thin()
            cell_v = ws2.cell(row=r, column=2, value=cnt)
            cell_v.font      = _font(color=fg, bold=True)
            cell_v.fill      = _fill(bg)
            cell_v.alignment = _align("center")
            cell_v.border    = _border_thin()
            for col in [3, 4]:
                _write_data_cell(ws2, r, col, "", False, bg=bg)
            r += 1

        r += 1
        _stat_header(ws2, r, "UMUMIY KO'RSATKICHLAR")
        r += 1
        total_det = sum(a.total_detections for a in animals)
        detected  = sum(1 for a in animals if a.total_detections > 0)
        gen_stats = [
            ("Jami jonivorlar",               len(animals)),
            ("Faol jonivorlar",               status_counts.get("active", 0)),
            ("Kuzatilgan jonivorlar",          detected),
            ("Kuzatilmagan jonivorlar",        len(animals) - detected),
            ("Jami deteksiyalar (umumiy)",     total_det),
            ("O'rtacha deteksiya/jonivor",     round(total_det / len(animals), 1) if animals else 0),
            ("Eksport sanasi",                 now_str),
        ]
        for i, (label, value) in enumerate(gen_stats):
            _stat_row(ws2, r, label, value, i % 2 == 0)
            r += 1

        _set_col_widths(ws2, [32, 16, 8, 8])

        # ------------------------------------------------------------------
        # 3. Workbook ni baytlarga aylantirish
        # ------------------------------------------------------------------
        buf = io.BytesIO()
        wb.save(buf)
        raw = buf.getvalue()
        buf.close()

        logger.info(f"[export] Animals Excel tayyor: {len(animals)} ta, {len(raw)} bayt")
        return raw

    # =========================================================================
    # DETECTIONS — CSV  (xatolar tuzatildi)
    # =========================================================================

    async def export_detections_csv(
        self,
        db: AsyncSession,
        date_from: date,
        date_to: date,
        animal_id: Optional[int] = None,
    ) -> bytes:
        """
        Deteksiyalarni CSV formatida eksport qiladi.

        Args:
            db:        DB sessiyasi
            date_from: Boshlanish sanasi
            date_to:   Tugash sanasi
            animal_id: Jonivor filtri (ixtiyoriy)

        Returns:
            UTF-8 kodlangan CSV baytlari

        CSV ustunlari:
            id, animal_id, animal_tag_id, camera_id,
            timestamp, confidence, class_name,
            bbox_x, bbox_y, bbox_w, bbox_h
        """
        logger.info(f"[export] Detections CSV: {date_from} → {date_to}, animal={animal_id}")

        # ✓ To'g'ri maydon nomi: Detection.timestamp (eski: detected_at)
        query = (
            select(Detection)
            .options(selectinload(Detection.animal))
            .where(
                and_(
                    func.date(Detection.timestamp) >= date_from,
                    func.date(Detection.timestamp) <= date_to,
                )
            )
        )
        if animal_id is not None:
            query = query.where(Detection.animal_id == animal_id)

        # ✓ To'g'ri tartiblash: Detection.timestamp (eski: detected_at)
        result     = await db.execute(query.order_by(Detection.timestamp.desc()))
        detections = result.scalars().all()

        rows = []
        for d in detections:
            bbox = d.bbox or {}
            rows.append({
                "id":            d.id,
                "animal_id":     d.animal_id,
                "animal_tag_id": d.animal.tag_id if d.animal else "Noma'lum",
                "camera_id":     d.camera_id,
                # ✓ To'g'ri: timestamp (eski: detected_at)
                "timestamp":     d.timestamp.isoformat(),
                # ✓ To'g'ri: confidence (eski: confidence_score)
                "confidence":    round(d.confidence, 4),
                "class_name":    d.class_name,
                # ✓ To'g'ri: bbox JSON (eski: bbox_x/y/width/height)
                "bbox_x":        round(bbox.get("x", 0), 4),
                "bbox_y":        round(bbox.get("y", 0), 4),
                "bbox_w":        round(bbox.get("w", 0), 4),
                "bbox_h":        round(bbox.get("h", 0), 4),
            })

        buf = io.BytesIO()
        pd.DataFrame(rows).to_csv(buf, index=False, encoding="utf-8")
        raw = buf.getvalue()
        buf.close()

        logger.info(f"[export] Detections CSV tayyor: {len(detections)} ta, {len(raw)} bayt")
        return raw

    # =========================================================================
    # WEIGHTS — EXCEL
    # =========================================================================

    async def export_weights_excel(
        self,
        db: AsyncSession,
        animal_ids: Optional[list[int]] = None,
    ) -> bytes:
        """
        Og'irlik o'lchovlarini ko'p varaqli Excel formatida eksport qiladi.

        EXCEL TARKIBI:
            Varaq 1 — "Xulosa": Barcha jonivornlar uchun umumiy
            Varaq 2+ — "Jonivor_{tag}": Har bir jonivor uchun alohida

        Args:
            db:         DB sessiyasi
            animal_ids: Jonivor IDlari filtri (ixtiyoriy, None = hammasi)

        Returns:
            .xlsx baytlari
        """
        logger.info(f"[export] Weights Excel: animal_ids={animal_ids}")

        animals_q = (
            select(Animal)
            .options(selectinload(Animal.weight_measurements))
            .where(Animal.status == AnimalStatus.ACTIVE)
        )
        if animal_ids:
            animals_q = animals_q.where(Animal.id.in_(animal_ids))

        result  = await db.execute(animals_q.order_by(Animal.tag_id))
        animals = result.scalars().all()

        buf    = io.BytesIO()
        writer = pd.ExcelWriter(buf, engine="openpyxl")

        # — XULOSA VARAG'I —
        summary_rows = []
        for a in animals:
            ms = a.weight_measurements
            if ms:
                wts    = [m.estimated_weight_kg for m in ms]
                latest = max(ms, key=lambda m: m.timestamp)
                summary_rows.append({
                    "Tag ID":              a.tag_id,
                    "Tur":                 _SPECIES_LABEL.get(
                                               a.species.value if hasattr(a.species, "value") else str(a.species),
                                               str(a.species)
                                           ),
                    "O'lchovlar soni":     len(ms),
                    "So'nggi og'irlik kg": round(latest.estimated_weight_kg, 2),
                    "So'nggi sana":        latest.timestamp.date().isoformat(),
                    "O'rtacha kg":         round(sum(wts) / len(wts), 2),
                    "Eng kam kg":          round(min(wts), 2),
                    "Eng ko'p kg":         round(max(wts), 2),
                })
            else:
                summary_rows.append({
                    "Tag ID":              a.tag_id,
                    "Tur":                 str(a.species),
                    "O'lchovlar soni":     0,
                    "So'nggi og'irlik kg": "—",
                    "So'nggi sana":        "—",
                    "O'rtacha kg":         "—",
                    "Eng kam kg":          "—",
                    "Eng ko'p kg":         "—",
                })

        pd.DataFrame(summary_rows).to_excel(writer, sheet_name="Xulosa", index=False)

        # — HAR JONIVOR UCHUN VARAQ —
        for a in animals:
            ms = a.weight_measurements
            if not ms:
                continue
            sorted_ms = sorted(ms, key=lambda m: m.timestamp, reverse=True)
            rows = [
                {
                    "Sana":        m.timestamp.date().isoformat(),
                    "Vaqt":        m.timestamp.strftime("%H:%M:%S"),
                    "Og'irlik kg": round(m.estimated_weight_kg, 2),
                    "Ishonch %":   round(m.confidence_score * 100, 1),
                    "Kamera":      m.camera_id,
                }
                for m in sorted_ms
            ]
            sheet_name = f"Jonivor_{a.tag_id}"[:31]
            pd.DataFrame(rows).to_excel(writer, sheet_name=sheet_name, index=False)

        writer.close()
        raw = buf.getvalue()
        buf.close()

        logger.info(f"[export] Weights Excel tayyor: {len(animals)} ta, {len(raw)} bayt")
        return raw

    # =========================================================================
    # TO'LIQ ARXIV — EXCEL  (xatolar tuzatildi)
    # =========================================================================

    async def export_all_data_excel(self, db: AsyncSession) -> bytes:
        """
        Barcha ferma ma'lumotlarini 4 varaqli Excel faylida eksport qiladi.

        EXCEL TARKIBI:
            Varaq 1 — "Jonivorlar"          : Hammasi
            Varaq 2 — "Deteksiyalar (30kun)": So'nggi 30 kun, max 10 000 qator
            Varaq 3 — "Ogʻirliklar"         : Barcha o'lchovlar, max 10 000
            Varaq 4 — "Statistika"          : Umumiy ko'rsatkichlar

        Returns:
            .xlsx baytlari
        """
        logger.info("[export] To'liq arxiv Excel boshlanmoqda")

        buf    = io.BytesIO()
        writer = pd.ExcelWriter(buf, engine="openpyxl")

        # — VARAQ 1: JONIVORLAR —
        animals_res = await db.execute(select(Animal).order_by(Animal.tag_id))
        animals     = animals_res.scalars().all()

        animals_rows = [
            {
                "ID":            a.id,
                "Tag ID":        a.tag_id,
                "Tur":           a.species.value if hasattr(a.species, "value") else str(a.species),
                "Jins":          a.gender.value  if hasattr(a.gender,  "value") else str(a.gender),
                "Holat":         a.status.value  if hasattr(a.status,  "value") else str(a.status),
                "Zot":           a.breed or "",
                "Sotib olingan": a.acquisition_date.date().isoformat() if a.acquisition_date else "",
                "Jami aniqlash": a.total_detections,
                "Oxirgi ko'rinish": a.last_detected_at.isoformat() if a.last_detected_at else "",
                "Izoh":          a.notes or "",
            }
            for a in animals
        ]
        pd.DataFrame(animals_rows).to_excel(writer, sheet_name="Jonivorlar", index=False)

        # — VARAQ 2: DETEKSIYALAR (so'nggi 30 kun) —
        # ✓ To'g'ri: Detection.timestamp (eski: Detection.detected_at)
        thirty_ago = datetime.now(timezone.utc).date() - timedelta(days=30)
        det_res = await db.execute(
            select(Detection)
            .options(selectinload(Detection.animal))
            .where(func.date(Detection.timestamp) >= thirty_ago)    # ← tuzatildi
            .order_by(Detection.timestamp.desc())                    # ← tuzatildi
            .limit(10_000)
        )
        detections = det_res.scalars().all()

        det_rows = [
            {
                "ID":            d.id,
                "Jonivor tag":   d.animal.tag_id if d.animal else "Noma'lum",
                "Kamera":        d.camera_id,
                # ✓ To'g'ri: timestamp (eski: detected_at)
                "Vaqt":          d.timestamp.isoformat(),
                # ✓ To'g'ri: confidence (eski: confidence_score)
                "Ishonch":       round(d.confidence, 4),
                "Klass":         d.class_name,
            }
            for d in detections
        ]
        pd.DataFrame(det_rows).to_excel(writer, sheet_name="Deteksiyalar (30kun)", index=False)

        # — VARAQ 3: OG'IRLIKLAR —
        wt_res = await db.execute(
            select(WeightMeasurement)
            .options(selectinload(WeightMeasurement.animal))
            .order_by(WeightMeasurement.timestamp.desc())
            .limit(10_000)
        )
        weights = wt_res.scalars().all()

        wt_rows = [
            {
                "ID":              w.id,
                "Jonivor tag":     w.animal.tag_id if w.animal else "Noma'lum",
                "Og'irlik kg":     round(w.estimated_weight_kg, 2),
                "Ishonch %":       round(w.confidence_score * 100, 1),
                "Vaqt":            w.timestamp.isoformat(),
                "Kamera":          w.camera_id,
            }
            for w in weights
        ]
        pd.DataFrame(wt_rows).to_excel(writer, sheet_name="Ogʻirliklar", index=False)

        # — VARAQ 4: STATISTIKA —
        active_count = sum(1 for a in animals if (a.status.value if hasattr(a.status, "value") else str(a.status)) == "active")
        stats_rows = [
            {"Ko'rsatkich": "Jami jonivorlar",               "Qiymat": len(animals)},
            {"Ko'rsatkich": "Faol jonivorlar",               "Qiymat": active_count},
            {"Ko'rsatkich": "Deteksiyalar (30 kun)",         "Qiymat": len(detections)},
            {"Ko'rsatkich": "Jami og'irlik o'lchovlari",     "Qiymat": len(weights)},
            {"Ko'rsatkich": "Eksport vaqti",                 "Qiymat": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")},
        ]
        pd.DataFrame(stats_rows).to_excel(writer, sheet_name="Statistika", index=False)

        writer.close()
        raw = buf.getvalue()
        buf.close()

        logger.info(f"[export] To'liq arxiv Excel tayyor: {len(raw)} bayt")
        return raw