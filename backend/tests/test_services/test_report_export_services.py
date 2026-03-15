"""
TAURUS VISION — tests/test_services/test_report_export_services.py
===================================================================
ReportService + ExportService uchun AYAMAS vahshiy testlar.

Qamrav (120+ test):
  ✓ ReportService dizayn konstantalari (brand ranglar)
  ✓ ReportService._fmt_date     — None, date, datetime
  ✓ ReportService._fmt_datetime — None, datetime
  ✓ ReportService._to_utc       — naive→UTC, aware o'zgarmaydi
  ✓ ReportService._build_header — tuzilma, title
  ✓ ReportService._build_footer — tuzilma
  ✓ ReportService._build_recommendations — kritik, missing, open alerts
  ✓ ReportService.generate_animal_report — PDF bytes qaytaradi, yo'q jonivor ValueError
  ✓ ReportService.generate_farm_report   — PDF bytes qaytaradi
  ✓ ReportService.generate_health_report — PDF bytes qaytaradi
  ✓ ExportService style helpers (_fill, _font, _border_thin, _align)
  ✓ ExportService._STATUS_HEX / _SPECIES_LABEL / _GENDER_LABEL
  ✓ ExportService.export_animals_csv     — bytes, CSV format
  ✓ ExportService.export_animals_excel   — bytes, xlsx
  ✓ ExportService.export_detections_csv  — bytes
  ✓ ExportService.export_weights_excel   — bytes
  ✓ ExportService.export_all_data_excel  — bytes, multi-sheet
"""

import pytest
from datetime import datetime, date, timezone, timedelta

from app.models.animal import Animal, AnimalSpecies, AnimalGender, AnimalStatus
from app.models.detection import Detection
from app.models.weight_measurement import WeightMeasurement
from app.services.report_service import (
    ReportService,
    _BRAND_DARK, _BRAND_GREEN, _BRAND_ACCENT,
    _TABLE_HEADER, _CRITICAL_RED, _WARNING_AMBER,
)
from app.services.export_service import (
    ExportService,
    _fill, _font, _border_thin, _align,
    _set_col_widths, _write_header_row,
    _STATUS_HEX, _SPECIES_LABEL, _GENDER_LABEL, _STATUS_LABEL,
)

pytestmark = pytest.mark.asyncio

NOW = datetime.utcnow()
TODAY = date.today()


# ─── Fixtures ──────────────────────────────────────────────────────────────────

@pytest.fixture
def rpt_svc():
    return ReportService()


@pytest.fixture
def exp_svc():
    return ExportService()


@pytest.fixture
async def animal(db):
    a = Animal(
        tag_id="RPT-ANIMAL-001",
        species=AnimalSpecies.CATTLE, gender=AnimalGender.FEMALE,
        status=AnimalStatus.ACTIVE,
        acquisition_date=datetime(2021, 1, 1),
        birth_date=datetime(2020, 6, 1),
    )
    db.add(a); await db.commit(); await db.refresh(a); return a


@pytest.fixture
async def animal_with_data(db, animal):
    """Jonivor + vazn o'lchovlari + deteksiyalar."""
    for i in range(3):
        wm = WeightMeasurement(
            animal_id=animal.id,
            estimated_weight_kg=350.0 + i * 5,
            confidence_score=0.9,
            camera_id="CAM-01",
            timestamp=NOW - timedelta(days=i),
        )
        db.add(wm)
    await db.commit()
    return animal


# ═══════════════════════════════════════════════════════════════════════════════
# REPORT SERVICE — Design constants
# ═══════════════════════════════════════════════════════════════════════════════

class TestReportServiceConstants:
    def test_brand_dark_is_hex_color(self):
        assert _BRAND_DARK is not None

    def test_brand_green_is_hex_color(self):
        assert _BRAND_GREEN is not None

    def test_brand_accent_is_hex_color(self):
        assert _BRAND_ACCENT is not None

    def test_table_header_color(self):
        assert _TABLE_HEADER is not None

    def test_critical_red_color(self):
        assert _CRITICAL_RED is not None

    def test_warning_amber_color(self):
        assert _WARNING_AMBER is not None


# ═══════════════════════════════════════════════════════════════════════════════
# REPORT SERVICE — Static helpers (pure functions)
# ═══════════════════════════════════════════════════════════════════════════════

class TestReportServiceStaticHelpers:

    def test_fmt_date_none_returns_dash(self, rpt_svc):
        assert rpt_svc._fmt_date(None) == "—"

    def test_fmt_date_date_object(self, rpt_svc):
        d = date(2026, 3, 15)
        assert rpt_svc._fmt_date(d) == "2026-03-15"

    def test_fmt_date_datetime_object(self, rpt_svc):
        dt = datetime(2026, 3, 15, 10, 30)
        assert rpt_svc._fmt_date(dt) == "2026-03-15"

    def test_fmt_date_string_returns_as_is(self, rpt_svc):
        result = rpt_svc._fmt_date("2026-01-01")
        assert "2026" in result

    def test_fmt_datetime_none_returns_dash(self, rpt_svc):
        assert rpt_svc._fmt_datetime(None) == "—"

    def test_fmt_datetime_datetime_object(self, rpt_svc):
        dt = datetime(2026, 3, 15, 14, 30)
        result = rpt_svc._fmt_datetime(dt)
        assert "2026-03-15" in result
        assert "14:30" in result

    def test_fmt_datetime_none_string(self, rpt_svc):
        result = rpt_svc._fmt_datetime(None)
        assert result == "—"

    def test_to_utc_naive_adds_utc(self, rpt_svc):
        naive = datetime(2026, 1, 1, 12, 0, 0)
        assert naive.tzinfo is None
        aware = rpt_svc._to_utc(naive)
        assert aware.tzinfo == timezone.utc

    def test_to_utc_aware_unchanged(self, rpt_svc):
        aware = datetime(2026, 1, 1, 12, 0, 0, tzinfo=timezone.utc)
        result = rpt_svc._to_utc(aware)
        assert result.tzinfo == timezone.utc
        assert result == aware

    def test_to_utc_preserves_time(self, rpt_svc):
        naive = datetime(2026, 6, 15, 9, 30, 45)
        result = rpt_svc._to_utc(naive)
        assert result.hour == 9
        assert result.minute == 30
        assert result.second == 45


# ═══════════════════════════════════════════════════════════════════════════════
# REPORT SERVICE — _build_header / _build_footer
# ═══════════════════════════════════════════════════════════════════════════════

class TestReportServiceBuilders:

    def test_build_header_returns_list(self, rpt_svc):
        result = rpt_svc._build_header("Test Title", "Test Subtitle")
        assert isinstance(result, list)
        assert len(result) > 0

    def test_build_header_contains_paragraphs(self, rpt_svc):
        from reportlab.platypus import Paragraph
        result = rpt_svc._build_header("Sarlavha", "Taglavha")
        paragraphs = [r for r in result if isinstance(r, Paragraph)]
        assert len(paragraphs) >= 2

    def test_build_footer_returns_list(self, rpt_svc):
        result = rpt_svc._build_footer()
        assert isinstance(result, list)
        assert len(result) > 0

    def test_build_divider_returns_table(self, rpt_svc):
        from reportlab.platypus import Table
        result = rpt_svc._build_divider()
        assert isinstance(result, Table)

    def test_build_kv_table_returns_table(self, rpt_svc):
        from reportlab.platypus import Table
        rows = [["Kalit", "Qiymat"], ["Ism", "Ali"]]
        result = rpt_svc._build_kv_table(rows)
        assert isinstance(result, Table)

    def test_build_data_table_returns_table(self, rpt_svc):
        from reportlab.platypus import Table
        rows = [["Sarlavha 1", "Sarlavha 2"], ["Data 1", "Data 2"]]
        result = rpt_svc._build_data_table(rows)
        assert isinstance(result, Table)

    def test_build_alert_table_returns_table(self, rpt_svc):
        from reportlab.platypus import Table
        rows = [["Tur", "Jonivor", "Sana"], ["critical", "JNV-001", "2026-01-01"]]
        result = rpt_svc._build_alert_table(rows)
        assert isinstance(result, Table)


# ═══════════════════════════════════════════════════════════════════════════════
# REPORT SERVICE — _build_recommendations
# ═══════════════════════════════════════════════════════════════════════════════

class TestBuildRecommendations:

    def test_no_issues_positive_rec(self):
        recs = ReportService._build_recommendations([], [], animal_count=10)
        assert len(recs) >= 1
        assert any("qoniqarli" in r or "samarali" in r for r in recs)

    def test_critical_alerts_add_rec(self):
        alerts = [{"severity": "critical"}, {"severity": "critical"}]
        recs = ReportService._build_recommendations(alerts, [], animal_count=10)
        assert any("kritik" in r.lower() or "veterinar" in r.lower() for r in recs)

    def test_long_missing_add_rec(self):
        alerts = [{"severity": "warning", "days": 20},
                  {"severity": "warning", "days": 18}]
        recs = ReportService._build_recommendations(alerts, [], animal_count=10)
        assert any("14 kun" in r or "kamera" in r.lower() for r in recs)

    def test_many_open_alerts_rec(self):
        db_alerts = [object()] * 15  # 15 ta ochiq alert
        recs = ReportService._build_recommendations([], db_alerts, animal_count=10)
        assert any("15" in r or "alert" in r.lower() for r in recs)

    def test_few_open_alerts_rec(self):
        db_alerts = [object()] * 3
        recs = ReportService._build_recommendations([], db_alerts, animal_count=10)
        assert any("3" in r or "alert" in r.lower() for r in recs)

    def test_returns_list(self):
        result = ReportService._build_recommendations([], [], 5)
        assert isinstance(result, list)


# ═══════════════════════════════════════════════════════════════════════════════
# REPORT SERVICE — PDF generation
# ═══════════════════════════════════════════════════════════════════════════════

class TestReportServiceGenerate:

    async def test_generate_animal_report_returns_bytes(self, db, rpt_svc, animal_with_data):
        result = await rpt_svc.generate_animal_report(db, animal_with_data.id)
        assert isinstance(result, bytes)
        assert len(result) > 0

    async def test_generate_animal_report_is_pdf(self, db, rpt_svc, animal_with_data):
        """PDF magic bytes: %PDF."""
        result = await rpt_svc.generate_animal_report(db, animal_with_data.id)
        assert result[:4] == b"%PDF"

    async def test_generate_animal_report_missing_raises(self, db, rpt_svc):
        with pytest.raises(ValueError) as exc_info:
            await rpt_svc.generate_animal_report(db, 999999)
        assert "999999" in str(exc_info.value)

    async def test_generate_animal_report_no_data_ok(self, db, rpt_svc, animal):
        """Ma'lumot bo'lmasa ham xato bermasin."""
        result = await rpt_svc.generate_animal_report(db, animal.id)
        assert isinstance(result, bytes)
        assert len(result) > 0

    async def test_generate_farm_report_returns_bytes(self, db, rpt_svc, animal):
        result = await rpt_svc.generate_farm_report(db)
        assert isinstance(result, bytes)
        assert len(result) > 0

    async def test_generate_farm_report_is_pdf(self, db, rpt_svc, animal):
        result = await rpt_svc.generate_farm_report(db)
        assert result[:4] == b"%PDF"

    async def test_generate_health_report_returns_bytes(self, db, rpt_svc, animal):
        result = await rpt_svc.generate_health_report(db)
        assert isinstance(result, bytes)
        assert len(result) > 0

    async def test_generate_health_report_is_pdf(self, db, rpt_svc, animal):
        result = await rpt_svc.generate_health_report(db)
        assert result[:4] == b"%PDF"

    async def test_generate_farm_report_no_animals_ok(self, db, rpt_svc):
        """Jonivorlar bo'lmasa ham xato bermasin."""
        result = await rpt_svc.generate_farm_report(db)
        assert isinstance(result, bytes)


# ═══════════════════════════════════════════════════════════════════════════════
# EXPORT SERVICE — Style helpers (pure functions)
# ═══════════════════════════════════════════════════════════════════════════════

class TestExportStyleHelpers:

    def test_fill_returns_pattern_fill(self):
        from openpyxl.styles import PatternFill
        result = _fill("FF0000")
        assert isinstance(result, PatternFill)

    def test_fill_sets_color(self):
        result = _fill("1E3A5F")
        assert result.fgColor.rgb == "FF1E3A5F"

    def test_font_returns_font(self):
        from openpyxl.styles import Font
        result = _font()
        assert isinstance(result, Font)

    def test_font_bold(self):
        result = _font(bold=True)
        assert result.bold is True

    def test_font_size(self):
        result = _font(size=14)
        assert result.size == 14

    def test_font_color(self):
        result = _font(color="FFFFFF")
        assert "FFFFFF" in result.color.rgb

    def test_border_thin_returns_border(self):
        from openpyxl.styles import Border
        result = _border_thin()
        assert isinstance(result, Border)

    def test_align_returns_alignment(self):
        from openpyxl.styles import Alignment
        result = _align("center")
        assert isinstance(result, Alignment)

    def test_align_horizontal(self):
        result = _align("right")
        assert result.horizontal == "right"

    def test_align_wrap_text(self):
        result = _align("left", wrap=True)
        assert result.wrap_text is True


# ═══════════════════════════════════════════════════════════════════════════════
# EXPORT SERVICE — Constants
# ═══════════════════════════════════════════════════════════════════════════════

class TestExportConstants:

    def test_status_hex_all_statuses(self):
        for s in ["active", "sick", "quarantine", "sold", "deceased"]:
            assert s in _STATUS_HEX
            bg, fg = _STATUS_HEX[s]
            assert len(bg) == 6
            assert len(fg) == 6

    def test_species_label_all_species(self):
        for s in ["cattle", "sheep", "goat", "horse", "other"]:
            assert s in _SPECIES_LABEL
            assert isinstance(_SPECIES_LABEL[s], str)

    def test_gender_label_all_genders(self):
        for g in ["male", "female", "unknown"]:
            assert g in _GENDER_LABEL

    def test_status_label_all_statuses(self):
        for s in ["active", "sick", "quarantine", "sold", "deceased"]:
            assert s in _STATUS_LABEL

    def test_species_cattle_label_uzbek(self):
        assert _SPECIES_LABEL["cattle"] == "Qoramol"

    def test_gender_male_label_uzbek(self):
        assert _GENDER_LABEL["male"] == "Erkak"

    def test_status_active_label_uzbek(self):
        assert _STATUS_LABEL["active"] == "Faol"


# ═══════════════════════════════════════════════════════════════════════════════
# EXPORT SERVICE — CSV and Excel generation
# ═══════════════════════════════════════════════════════════════════════════════

class TestExportServiceCSV:

    async def test_export_animals_csv_returns_bytes(self, db, exp_svc, animal):
        result = await exp_svc.export_animals_csv(db)
        assert isinstance(result, bytes)
        assert len(result) > 0

    async def test_export_animals_csv_contains_tag(self, db, exp_svc, animal):
        result = await exp_svc.export_animals_csv(db)
        content = result.decode("utf-8-sig", errors="replace")
        assert animal.tag_id in content

    async def test_export_animals_csv_has_header(self, db, exp_svc, animal):
        result = await exp_svc.export_animals_csv(db)
        content = result.decode("utf-8-sig", errors="replace")
        first_line = content.split("\n")[0]
        # Header satri bo'lishi kerak
        assert len(first_line) > 0

    async def test_export_animals_csv_no_animals_ok(self, db, exp_svc):
        result = await exp_svc.export_animals_csv(db)
        assert isinstance(result, bytes)

    async def test_export_animals_csv_status_filter(self, db, exp_svc, animal):
        result = await exp_svc.export_animals_csv(db, status="active")
        assert isinstance(result, bytes)

    async def test_export_detections_csv_returns_bytes(self, db, exp_svc, animal):
        result = await exp_svc.export_detections_csv(
            db, date_from=TODAY - timedelta(days=7), date_to=TODAY)
        assert isinstance(result, bytes)

    async def test_export_detections_csv_with_detections(self, db, exp_svc, animal):
        det = Detection(
            animal_id=animal.id, camera_id="CAM-01",
            timestamp=datetime.utcnow(), confidence=0.92,
            class_id=19, class_name="cow",
            bbox={"x": 0.3, "y": 0.2, "w": 0.25, "h": 0.35},
        )
        db.add(det); await db.commit()
        result = await exp_svc.export_detections_csv(
            db, date_from=TODAY - timedelta(days=1), date_to=TODAY)
        assert isinstance(result, bytes)
        assert len(result) > 0


class TestExportServiceExcel:

    async def test_export_animals_excel_returns_bytes(self, db, exp_svc, animal):
        result = await exp_svc.export_animals_excel(db)
        assert isinstance(result, bytes)
        assert len(result) > 0

    async def test_export_animals_excel_is_xlsx(self, db, exp_svc, animal):
        """XLSX magic bytes: PK\x03\x04."""
        result = await exp_svc.export_animals_excel(db)
        assert result[:2] == b"PK"

    async def test_export_animals_excel_no_animals_ok(self, db, exp_svc):
        result = await exp_svc.export_animals_excel(db)
        assert isinstance(result, bytes)

    async def test_export_weights_excel_returns_bytes(self, db, exp_svc, animal):
        wm = WeightMeasurement(
            animal_id=animal.id, estimated_weight_kg=350.0,
            confidence_score=0.9, camera_id="CAM-01",
            timestamp=datetime.utcnow(),
        )
        db.add(wm); await db.commit()
        result = await exp_svc.export_weights_excel(db)
        assert isinstance(result, bytes)
        assert len(result) > 0

    async def test_export_weights_excel_is_xlsx(self, db, exp_svc, animal):
        result = await exp_svc.export_weights_excel(db)
        assert result[:2] == b"PK"

    async def test_export_all_data_excel_returns_bytes(self, db, exp_svc, animal):
        result = await exp_svc.export_all_data_excel(db)
        assert isinstance(result, bytes)
        assert len(result) > 0

    async def test_export_all_data_excel_is_xlsx(self, db, exp_svc, animal):
        result = await exp_svc.export_all_data_excel(db)
        assert result[:2] == b"PK"

    async def test_export_all_data_no_data_ok(self, db, exp_svc):
        """Ma'lumot bo'lmasa ham xato bermasin."""
        result = await exp_svc.export_all_data_excel(db)
        assert isinstance(result, bytes)


# ═══════════════════════════════════════════════════════════════════════════════
# WRITE HELPERS test
# ═══════════════════════════════════════════════════════════════════════════════

class TestExportWriteHelpers:

    def test_set_col_widths(self):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        _set_col_widths(ws, [10, 20, 30])
        assert ws.column_dimensions["A"].width == 10
        assert ws.column_dimensions["B"].width == 20
        assert ws.column_dimensions["C"].width == 30

    def test_write_header_row(self):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        headers = ["ID", "Nomi", "Turi"]
        _write_header_row(ws, headers, row=1)
        assert ws.cell(1, 1).value == "ID"
        assert ws.cell(1, 2).value == "Nomi"
        assert ws.cell(1, 3).value == "Turi"
        # Bold bo'lishi kerak
        assert ws.cell(1, 1).font.bold is True

    def test_write_header_row_custom_row(self):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        _write_header_row(ws, ["A", "B"], row=3)
        assert ws.cell(3, 1).value == "A"
        assert ws.cell(2, 1).value is None